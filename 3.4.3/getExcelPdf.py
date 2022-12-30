import os
import numpy as np
import pandas as pd
import openpyxl
from matplotlib import pyplot as plt
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from jinja2 import Environment, FileSystemLoader
import pdfkit
from xlsx2html import xlsx2html



class PdfMaker:
    """Класс для конвертирования данных статистики в pdf-файл"""

    def __init__(self, graph_name, excel_file_name, profession):
        """Инициализирует PdfMaker"""
        self.graph = graph_name
        self.excel_file = excel_file_name
        self.prof = profession

    def generate_pdf(self):
        """Генерирует pdf файл"""
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")
        graph_path = os.path.abspath(self.graph)
        out_stream = xlsx2html(self.excel_file, sheet="Статистика по годам")
        out_stream.seek(0)
        pdf_template = template.render({"profession": self.prof, "graph-image": graph_path, "table": out_stream.read()})
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={"enable-local-file-access": ""})


class ExcelMaker:
    """Класс для создания Excel"""

    def __init__(self, profession, years, average_salary, average_salary_profession, count_vacancies_by_year,
                 count_vacancies_by_year_prof, file_name):
        """Инициализирует ExcelMaker"""
        self.years = years
        self.average_salary = average_salary
        self.average_salary_profession = average_salary_profession
        self.count_vacancies_by_year = count_vacancies_by_year
        self.count_vacancies_by_year_prof = count_vacancies_by_year_prof
        self.profession = profession
        self.file_name = file_name

    def generate_excel(self):
        """Генерирует excel файл"""
        if not isinstance(self.file_name, str) \
                or os.path.basename(self.file_name).split('.')[1] != "xlsx" \
                or os.path.exists(self.file_name):
            raise TypeError('')
        df = [
            [self.years[i], self.average_salary[i], self.average_salary_profession[i], self.count_vacancies_by_year[i],
             self.count_vacancies_by_year_prof[i]] for i in range(len(self.years))]
        df.insert(0, ["Год", "Средняя зарплата", f"Средняя зарплата - {self.profession}", "Количество вакансий",
                      f"Количество вакансий - {self.profession}"])
        df = pd.DataFrame(df, columns=None)
        with pd.ExcelWriter(self.file_name) as writer:
            df.to_excel(writer, sheet_name='Статистика по годам', index=False, header=False)
        wb = openpyxl.load_workbook(self.file_name)
        worksheet1 = wb["Статистика по годам"]
        thin = Side(border_style="thin")
        self.add_border(worksheet1, thin, len(self.years) + 2, ["A", "B", "C", "D", "E"])
        self.settle_width(worksheet1)
        wb.save(self.file_name)

    def add_border(self, worksheet, side, count_columns, rows):
        """Добавляет необходимым ячейкам обводку"""
        for i in range(1, count_columns):
            for row in rows:
                if i < 2:
                    worksheet[row + str(i)].alignment = Alignment(horizontal='left')
                    worksheet[row + str(i)].font = Font(bold=True)
                if worksheet[row + str(i)].internal_value is not None:
                    worksheet[row + str(i)].border = Border(top=side, bottom=side, left=side, right=side)

    def settle_width(self, worksheet):
        """Добавляет необходимым ячейкам ширину"""
        dims = {}
        for row in worksheet.rows:
            for cell in row:
                if cell.value is not None:
                    dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
                else:
                    dims[cell.column] = len(str(cell.value))
        for col, value in dims.items():
            worksheet.column_dimensions[get_column_letter(col)].width = value + 2


class StatisticFilesMaker:
    """ Класс для создания итоговых файлов"""

    def __init__(self, year_salary, year_vacancy, professions_year_salary, professions_year_vacancies, profession):
        """Инициализирует StatisticFilesMaker"""
        self.year_salary = year_salary
        self.year_vacancy = year_vacancy
        self.professions_year_salary = professions_year_salary
        self.professions_year_vacancies = professions_year_vacancies
        self.profession = profession

    def create_files(self):
        """Создает pdf с общей информацией"""
        output_data = {"Динамика уровня зарплат по годам:": self.year_salary,
                       "Динамика количества вакансий по годам:": self.year_vacancy,
                       "Динамика уровня зарплат по годам для выбранной профессии:": self.professions_year_salary,
                       "Динамика количества вакансий по годам для выбранной профессии:": self.professions_year_vacancies}
        [print(i, output_data[i]) for i in output_data]
        report = ExcelMaker(profession=self.profession,
                            years=[i for i in self.year_salary],
                            average_salary=[self.year_salary[i] for i in self.year_salary],
                            average_salary_profession=[self.professions_year_salary[i] for i in self.professions_year_salary],
                            count_vacancies_by_year=[self.year_vacancy[i] for i in self.year_vacancy],
                            count_vacancies_by_year_prof=[self.professions_year_vacancies[i] for i in self.professions_year_vacancies],
                            file_name="report.xlsx")
        report.generate_excel()
        MakeGraph(profession=self.profession,
                years=[i for i in self.year_salary],
                average_salary=[self.year_salary[i] for i in self.year_salary],
                average_salary_profession=[self.professions_year_salary[i] for i in self.professions_year_salary],
                count_vacancies_by_year=[self.year_vacancy[i] for i in self.year_vacancy],
                count_vacancies_by_year_prof=[self.professions_year_vacancies[i] for i in
                                              self.professions_year_vacancies],
                file_name="graph.png")
        pdf = PdfMaker(graph_name="graph.png", excel_file_name="report.xlsx", profession=self.profession)
        pdf.generate_pdf()


class MakeGraph:
    """Класс для создания графиков"""
    def __init__(self, profession, years, average_salary, average_salary_profession, count_vacancies_by_year,
                 count_vacancies_by_year_prof, file_name):
        """Инициализирует MakeGraph"""
        if not isinstance(file_name, str) \
                or os.path.basename(file_name).split('.')[1] != "png" \
                or os.path.exists(file_name):
            raise TypeError('')
        self.years = years
        self.average_salary = average_salary
        self.average_salary_profession = average_salary_profession
        self.count_vacancies_by_year = count_vacancies_by_year
        self.count_vacancies_by_year_prof = count_vacancies_by_year_prof
        self.profession = profession
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 8))
        self.generate_graph(ax1, "Уровень зарплат по годам", self.average_salary, self.years,
                            self.average_salary_profession, 'средняя з/п', f'з/п {self.profession}')
        self.generate_graph(ax2, 'Количество вакансий по годам', self.count_vacancies_by_year, self.years,
                            self.count_vacancies_by_year_prof, 'Количество вакансий', f'Количество вакансий {self.profession}')
        plt.tight_layout()
        fig.savefig(file_name)

    def generate_graph(self, ax, title, values_x, values_y, values_x2, label_x, label_x2):
        """Генерирует графики по переданной информации
        """
        ax.grid(axis='y')
        x = np.arange(len(values_y))
        width = 0.4
        ax.bar(x - width / 2, values_x, width, label=label_x)
        ax.bar(x + width / 2, values_x2, width, label=label_x2)
        ax.set_xticks(x, values_y, rotation=90)
        ax.tick_params(axis="both", labelsize=16)
        ax.set_title(title, fontweight='normal', fontsize=20)
        ax.legend(loc="upper left", fontsize=14)