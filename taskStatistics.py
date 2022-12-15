import csv
import datetime
from operator import itemgetter
import openpyxl
import numpy as np
import matplotlib.pyplot as plt
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from jinja2 import Environment, FileSystemLoader
import pdfkit
import doctest

currency_to_rub = {"AZN": 35.68,
                   "BYR": 23.91,
                   "EUR": 59.90,
                   "GEL": 21.74,
                   "KGS": 0.76,
                   "KZT": 0.13,
                   "RUR": 1,
                   "UAH": 1.64,
                   "USD": 60.66,
                   "UZS": 0.0055, }


def take_first_ten(dictionary):
    # """
    #     Функция возвращающая первые десять элемнтов словаря
    #     Parameters:
    #                dictionary (dict): словарь
    #     Returns:
    #               (dict): новый словарь из 10 элементов
    #     >>> take_first_ten({'Казань': 156337, 'Москва': 142291, 'Санкт-Петербург': 111548, 'Уфа': 106750, 'Екатеринбург': 95270, 'Владивосток': 87916, 'Набережные Челны': 81142, 'Иркутск': 80357, 'Нижний Новгород': 74437, 'Краснодар': 70402, 'Ростов-на-Дону': 68961, 'Хабаровск': 62800, 'Алматы': 61152, 'Тюмень': 59900, 'Ижевск': 58200, 'Красноярск': 57833, 'Новосибирск': 56958, 'Пермь': 55888, 'Челябинск': 52402, 'Томск': 46225, 'Минск': 45560})
    #     {'Казань': 156337, 'Москва': 142291, 'Санкт-Петербург': 111548, 'Уфа': 106750, 'Екатеринбург': 95270, 'Владивосток': 87916, 'Набережные Челны': 81142, 'Иркутск': 80357, 'Нижний Новгород': 74437, 'Краснодар': 70402}
    #
    #     >>> take_first_ten({'Казань': 156337, 'Москва': 142291, 'Санкт-Петербург': 111548, 'Уфа': 106750})
    #     {'Казань': 156337, 'Москва': 142291, 'Санкт-Петербург': 111548, 'Уфа': 106750}
    #
    #     >>> take_first_ten({})
    #     {}
    # """
    new_dictionary = {}
    i = 0
    for key in dictionary:
        new_dictionary[key] = round(dictionary[key], 4)
        i += 1
        if i == 10: break
    return new_dictionary

# def format_year_1(date):
#     """Функция для форматирования года
#         Args:
#             date (str): дата в виде строки
#         Returns:
#             str: отформатированная дата
#     """
#     date = datetime.datetime.strptime(date, "%Y-%m-%dT%H:%M:%S%z")
#     return date.year


# def format_year_2(date):
#     """Функция для форматирования года
#         Args:
#             date (str): дата в виде строки
#         Returns:
#             str: отформатированная дата
#     """
#     date = date.split('-')[0]
#     return date

def format_year_3(date):
    """Функция для форматирования года
        Args:
            date (str): дата в виде строки
        Returns:
            str: отформатированная дата
    """
    return date[:4]

class Vacancy:
    '''
        Класс для предоставления вакансии
        Attributes:
            dictionary (dict): словарь
        '''

    def __init__(self, dictionary):
        # '''
        #     Инициализирует объект вакансии. Выполняет структурирование информации о вакансии
        #     Attributes:
        #             dictionary (dict): словарь с данными о вакансии
        #
        #     >>> Vacancy({'name': '','salary_from': '26000.0', 'salary_to': '35000.0','salary_currency': 'RUR', 'area_name': '', 'published_at': '2022'}).salary
        #     30500.0
        #
        #     >>> Vacancy({'name': '','salary_from': '0', 'salary_to': '0','salary_currency': 'RUR', 'area_name': '', 'published_at': '2022'}).published_at
        #     2022
        #
        #     >>> Vacancy({'name': '','salary_from': '0', 'salary_to': '0','salary_currency': 'RUR', 'area_name': 'Краснотурьинск', 'published_at': '2022'}).area_name
        #     'Краснотурьинск'
        #
        #     >>> Vacancy({'name': 'Техник по связи г. Краснотурьинск','salary_from': '0', 'salary_to': '0','salary_currency': 'RUR', 'area_name': '', 'published_at': '2022'}).name
        #     'Техник по связи г. Краснотурьинск'
        # '''
        self.name = dictionary["name"]
        self.salary = (float(dictionary["salary_from"]) + float(dictionary["salary_to"])) / 2 * currency_to_rub[
            dictionary["salary_currency"]]
        self.area_name = dictionary["area_name"]
        self.published_at = int(format_year_3(dictionary["published_at"]))
        # self.published_at = int(dictionary["published_at"][:4])


class DataSet:
    """Считывание файла и формирование структуры данных о нем.
           Attributes:
               name (str): название csv файла.
               profession (str): Название профессии.
       """

    def __init__(self, name, profession):
        """Инициализирует класс DataSet. Чтение фала,форматирование,вывод информации.
                       Args:
                           name (str): название csv файла.
                           profession (str): название профессии.
               """
        self.file_name = name
        self.profession = profession
        headings, vacancies = self.csv_reader()
        dictionaries = self.csv_filer(vacancies, headings)
        self.vacancies_objects = [Vacancy(dictionary) for dictionary in dictionaries]
        vacancies_dicts = self.get_vacancies_into_dicts()
        self.vacancies_amount_at_times = DataSet.get_vacancies_amount_at_times(self.get_publised_list())
        self.vacancies_amount_at_times_for_profession = self.get_vacancies_amount_at_times_for_profession(vacancies_dicts)
        self.salary_at_times = self.get_salary_at_times()
        self.salary_at_times_for_profession = self.get_salary_at_times_for_profession()
        self.vacancies_amount_in_cities = self.get_vacancies_amount_in_cities(vacancies_dicts)
        self.vacancies_share_in_cities = self.get_vacancies_share_in_cities()
        self.salary_in_cities = self.get_salary_in_cities()

    def csv_reader(self):
        '''Чтение данных из csv файла
                Returns:
                    headings (list): список из заголовков фалйла
                    vacancies (list): список из профессий
        '''
        headings = []
        vacancies = []
        columns = 0
        rows = 0
        is_headings = True
        with open(self.file_name, encoding="utf-8-sig") as File:
            reader = csv.reader(File)
            for row in reader:
                rows += 1
                if is_headings:
                    headings = row
                    columns = len(row)
                    is_headings = False
                else:
                    if "" not in row and len(row) == columns:
                        vacancies.append(row)
                    else:
                        continue
        if rows < 2:
            print("Пустой файл" if rows == 0 else "Нет данных")
            exit()
        return headings, vacancies

    def csv_filer(self, reader, headings):
        """Записывает в словари данные полученные после чтения csv файла
        Args:
            reader (list): список вакансий полученных после обработки csv файла
            headings (list): список заголовков полученных после обработки csv файла
        Returns:
            dictionaries (list[dict]): список словарей с вакансиями
        """
        dictionaries = []
        for vacancy in reader:
            dictionary = {}
            for i in range(len(headings)):
                dictionary[headings[i]] = vacancy[i]
            dictionaries.append(dictionary)
        return dictionaries

    def get_salary_at_times(self):
        """Вычисляет среднее значение зарплаты относительно года
           Returns:
               (dict): отсоритованный относительно года словарь с зарплатами

        """
        dictionary = {}
        for vacancy in self.vacancies_objects:
            if vacancy.published_at not in dictionary:
                dictionary[vacancy.published_at] = vacancy.salary
            else:
                dictionary[vacancy.published_at] += vacancy.salary
        for key in dictionary:
            dictionary[key] = int(dictionary[key] / self.vacancies_amount_at_times[key])
        return dict(sorted(dictionary.items(), key=itemgetter(0)))

    def get_publised_list(self):
        """Возвращает лист с датами публикаций

           Returns:
               (list): лист с датами публикаций
        """
        return [vacancy.published_at for vacancy in self.vacancies_objects]

    def get_vacancies_into_dicts(self):
        """Преобразует объекты выкансии в словари

        (list): лист с словарями-вакансиями
        """
        vacancies_list = []
        for vacancy in self.vacancies_objects:
            dict = {}
            dict["name"] = vacancy.name
            dict["salary"] = vacancy.salary
            dict["area_name"] = vacancy.area_name
            dict["published_at"] = vacancy.published_at
            vacancies_list.append(dict)
        return vacancies_list

    @staticmethod
    def get_vacancies_amount_at_times(publised_list):
        # """Вычисляет кол-во вакансий относительно года
        #     Args:
        #         publised_list (list): лист с датами публикаций
        #     Returns:
        #               (dict): отсоритованный относительно года словарь с кол-вом вакансий
        #    >>> DataSet.get_vacancies_amount_at_times([2022,2019,2022,2021])
        #    {2019: 1, 2021: 1, 2022: 2}
        #
        #    >>> DataSet.get_vacancies_amount_at_times([2022,2022,2022,2022])
        #    {2022: 4}
        #
        #    >>> DataSet.get_vacancies_amount_at_times([2007,2010,2007,2001,2005,1992,1999])
        #    {1992: 1, 1999: 1, 2001: 1, 2005: 1, 2007: 2, 2010: 1}
        #
        #    >>> DataSet.get_vacancies_amount_at_times([])
        #    {}
        # """
        dictionary = {}
        for data in publised_list:
            if data in dictionary:
                dictionary[data] += 1
            else:
                dictionary[data] = 1
        return dict(sorted(dictionary.items(), key=itemgetter(0)))

    def get_salary_at_times_for_profession(self):
        """Вычисляет среднее значение зарплаты относительно года по выбранной профессии
           Returns:
             (dict): отсоритованный относительно года словарь с зарплатами по выбранной профессии
        """
        dictionary = {}
        for vacancy in self.vacancies_objects:
            if self.profession not in vacancy.name:
                continue
            if vacancy.published_at not in dictionary:
                dictionary[vacancy.published_at] = vacancy.salary
            else:
                dictionary[vacancy.published_at] += vacancy.salary
        for key in dictionary:
            dictionary[key] = int(dictionary[key] / self.vacancies_amount_at_times_for_profession[key])
        dictionary = dict(sorted(dictionary.items(), key=itemgetter(0)))
        return dictionary if len(dictionary) != 0 else {2022: 0}

    def get_vacancies_amount_at_times_for_profession(self, list_vacancies):
        # """Вычисляет кол-во вакансий относительно года по выбранной профессии
        #     Args:
        #         list_vacancies (list): лист с объектами-словарями
        #
        #    Returns:
        #       (dict): отсоритованный относительно года словарь с кол-вом вакансий по выбранной профессии
        #
        #    >>> dicts = [{'name': 'Механик','published_at': 2022},{'name': 'Аналитик','published_at': 2022} ,{'name': 'Аналитик','published_at': 2020}]
        #    >>> DataSet('vacancies.csv','Аналитик').get_vacancies_amount_at_times_for_profession(dicts)
        #    {2020: 1, 2022: 1}
        #
        #    >>> dicts = [{'name': 'Аналитик','published_at': 2022},{'name': 'Аналитик','published_at': 2022} ,{'name': 'Аналитик','published_at': 2020}]
        #    >>> DataSet('vacancies.csv','Аналитик').get_vacancies_amount_at_times_for_profession(dicts)
        #    {2020: 1, 2022: 2}
        #
        #    >>> dicts = []
        #    >>> DataSet('vacancies.csv','Аналитик').get_vacancies_amount_at_times_for_profession(dicts)
        #    {2022: 0}
        # """
        dictionary = {}
        for vacancy in list_vacancies:
            if self.profession not in vacancy['name']:
                continue
            if vacancy['published_at'] in dictionary:
                dictionary[vacancy['published_at']] += 1
            else:
                dictionary[vacancy['published_at']] = 1
        dictionary = dict(sorted(dictionary.items(), key=itemgetter(0)))
        return dictionary if len(dictionary) != 0 else {2022: 0}

    def get_vacancies_amount_in_cities(self, vacancies_dicts):
        # """Вычисляет кол-во вакансий по городам
        #    Args:
        #         vacancies_dicts (list): лист с объектами-словарями
        #    Returns:
        #            (dict): словарь с кол-вом вакансий по городам
        #    >>> dicts = [{'area_name': 'МСК'},{'area_name': 'МСК'},{'area_name': 'МСК'}]
        #    >>> DataSet('vacancies.csv', 'Аналитик').get_vacancies_amount_in_cities(dicts)
        #    {'МСК': 3}
        #
        #    >>> dicts = [{'area_name': 'МСК'},{'area_name': 'СПБ'},{'area_name': 'СПБ'}]
        #    >>> DataSet('vacancies.csv', 'Аналитик').get_vacancies_amount_in_cities(dicts)
        #    {'МСК': 1, 'СПБ': 2}
        #
        #    >>> dicts = [{'area_name': 'МСК'},{'area_name': 'ЕКБ'},{'area_name': 'СПБ'}]
        #    >>> DataSet('vacancies.csv', 'Аналитик').get_vacancies_amount_in_cities(dicts)
        #    {'МСК': 1, 'ЕКБ': 1, 'СПБ': 1}
        #
        #    >>> dicts = []
        #    >>> DataSet('vacancies.csv', 'Аналитик').get_vacancies_amount_in_cities(dicts)
        #    {}
        # """
        dictionary = {}
        for vacancy in vacancies_dicts:
            if vacancy['area_name'] not in dictionary:
                dictionary[vacancy['area_name']] = 1
            else:
                dictionary[vacancy['area_name']] += 1
        return dictionary

    def get_salary_in_cities(self):
        """Вычисляет величину зарплат по городам
           Returns:
              (dict): словарь из первых 10 значений с величинами зарплат по городам
        """
        dictionary = {}
        for vacancy in self.vacancies_objects:
            if self.vacancies_amount_in_cities[vacancy.area_name] / len(self.vacancies_objects) < 0.01:
                continue
            if vacancy.area_name not in dictionary:
                dictionary[vacancy.area_name] = vacancy.salary
            else:
                dictionary[vacancy.area_name] += vacancy.salary
        for key in dictionary:
            dictionary[key] = int(dictionary[key] / self.vacancies_amount_in_cities[key])
        return take_first_ten(dict(sorted(dictionary.items(), key=itemgetter(1), reverse=True)))

    def get_vacancies_share_in_cities(self):
        """Вычисляет города в которых кол-во вакансий больше или равно 1% от общего числа вакансий
            Returns:
              (dict): словарь из первых 10 вакансий кол-во которых больше или равно 1% от общего числа вакансий
        """
        dictionary = {}
        for key in self.vacancies_amount_in_cities:
            if self.vacancies_amount_in_cities[key] / len(self.vacancies_objects) >= 0.01:
                dictionary[key] = self.vacancies_amount_in_cities[key] / len(self.vacancies_objects)
        return take_first_ten(dict(sorted(dictionary.items(), key=itemgetter(1), reverse=True)))

    def print_result(self):
        """Выводит данные необходимы для графиков в консоль
        """
        print("Динамика уровня зарплат по годам: " + str(self.salary_at_times))
        print("Динамика количества вакансий по годам: " + str(self.vacancies_amount_at_times))
        print("Динамика уровня зарплат по годам для выбранной профессии: " + str(
            self.salary_at_times_for_profession))
        print("Динамика количества вакансий по годам для выбранной профессии: " + str(
            self.vacancies_amount_at_times_for_profession))
        print("Уровень зарплат по городам (в порядке убывания): " + str(self.salary_in_cities))
        print("Доля вакансий по городам (в порядке убывания): " + str(self.vacancies_share_in_cities))


class report:
    """Класс создающий  png-график, excel таблицу и pdf-файл с таблицей и графиком.
       Attributes:
           dataset (DataSet): объект с данными необходимыми для построения графиков и таблиц
       """

    def __init__(self, dataset):
        """Инициализирует класс report. Структурирование данных для графиков и таблиц.
                Args:
                    data (DataSet): Посчитанные данные для графиков.
        """
        self.profession = dataset.profession
        self.years_list_headings = (
            "Год", "Средняя зарплата", f"Средняя зарплата - {self.profession}", "Количество вакансий",
            f"Количество вакансий - {self.profession}")
        self.cities_list_headings = ("Город", "Уровень зарплат", "", "Город", "Доля вакансий")
        self.years_list_items = [[year for year in dataset.salary_at_times],
                                 [value for value in dataset.salary_at_times.values()],
                                 [value for value in dataset.salary_at_times_for_profession.values()],
                                 [value for value in dataset.vacancies_amount_at_times.values()],
                                 [value for value in dataset.vacancies_amount_at_times_for_profession.values()]]
        self.cities_list_items = [[city for city in dataset.salary_in_cities],
                                  [value for value in dataset.salary_in_cities.values()],
                                  ["" for i in range(len(dataset.salary_in_cities))],
                                  [city for city in dataset.vacancies_share_in_cities],
                                  [value for value in dataset.vacancies_share_in_cities.values()]]
        self.years_list_widths = [len(heading) + 2 for heading in self.years_list_headings]
        self.cities_list_widths = [len(heading) + 2 for heading in self.cities_list_headings]
        self.set_widths_xl(self.years_list_items, self.years_list_widths)
        self.set_widths_xl(self.cities_list_items, self.cities_list_widths)

    def set_widths_xl(self, column_items, column_widths):
        """Устанавливает ширину для каждой колонки по максимально широкому элементу
           Args:
               column_items(list): элементы колонок
               column_widths(list): лист заголовков для колонок
        """
        for i in range(len(column_items)):
            for item in column_items[i]:
                column_widths[i] = max(len(str(item)) + 2, column_widths[i])

    def clean_column(self, list_items, name):
        """Очищает стили колонки
            Args:
                name(str): название колонки для очищения из excel листа
                list_items (list): excel лист
        """
        for cell in list_items[name]:
            cell.border = Border(top=Side(border_style=None), bottom=Side(border_style=None))

    def make_border(self, list_items, width, height):
        """Создает границы для ячеек таблицы
            Args:
                list_items (list): excel лист
                width: ширина таблицы
                height: высота таблицы
        """
        cell_range = f'A1:{get_column_letter(width)}{height}'
        thin = Side(border_style="thin", color="000000")
        for row in list_items[cell_range]:
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    def set_bold_cells(self, cells):
        """Устанавливает толщину шрифта ячеек как bold
             Args:
                  cells(list) : лист с ячейками
        """
        for cell in cells:
            cell.font = Font(bold=True)

    def fill_rows(self, list, items, column_height):
        """Заполняет ряды таблицы
            Args:
                list (list): excel лист
                items (list[list]): элементы ячеек колонок
                column_height (int): высота колонки
        """
        for i in range(column_height):
            list.append([column[i] for column in items])

    def set_format(self, list_column, name, format):
        """Устанавливает формат ячеек колонки excel листа
            Args:
                list_sheet (list): excel лист
                name (str): наименование колонки таблицы
                format (str): устанавливаемый формат для ячеек
        """
        for cell in list_column[name]:
            cell.number_format = format

    def adjust_column_size(self, list, widths):
        """Устанавливает ширины для excel листа по самым широким элементам
        Args:
           list_sheet (list): excel лист
           widths (list[int]): ширины самых длинных элементов для каждой колонки
        """
        for i in range(1, 6):
            list.column_dimensions[get_column_letter(i)].width = widths[i - 1]

    def generate_excel(self):
        """Генерирует excel файл с двумя страницами о вакансиях и таблицами в каждой:
            по годам и по городам соотвественно
           Returns:
               (list): лист со станицами по годам и по городам
        """
        wb = openpyxl.Workbook()
        years_list = wb.active
        years_list.title = "Статистика по годам"
        cities_list = wb.create_sheet("Статистика по городам")
        years_list.append(self.years_list_headings)
        cities_list.append(self.cities_list_headings)
        self.set_bold_cells(years_list['1'])
        self.set_bold_cells(cities_list['1'])
        self.fill_rows(years_list, self.years_list_items, len(self.years_list_items[0]))
        self.fill_rows(cities_list, self.cities_list_items, len(self.cities_list_items[0]))
        self.set_format(cities_list, 'E', FORMAT_PERCENTAGE_00)
        self.adjust_column_size(cities_list, self.cities_list_widths)
        self.adjust_column_size(years_list, self.years_list_widths)
        self.make_border(years_list, len(self.years_list_headings), len(self.years_list_items[0]) + 1)
        self.make_border(cities_list, len(self.cities_list_headings), len(self.cities_list_items[0]) + 1)
        self.clean_column(cities_list, 'C')
        wb.save('report.xlsx')
        return [years_list, cities_list]

    def generate_vertical_graph(self, title, legend_titles, param1, param2, labels, ax):
        """Генерирует вертикальный график
                Args:
                   title (str): название графика
                   legend_titles (list[str]): лист с названиями для легенд
                   param1 (list): первый парметр для отображения на графике
                   param2 (list): второй парметр для отображения на графике
                   labels (list): лист c элементами для оси OX
                   ax: область в поле для рисования графика
            """
        x = np.arange(len(labels))
        width = 0.35
        plt.rcParams['font.size'] = '8'
        ax.bar(x - width / 2, param1, width, label=legend_titles[0])
        ax.bar(x + width / 2, param2, width, label=legend_titles[1])
        ax.set_xticks(x, labels, rotation="vertical")
        ax.grid(axis='y')
        ax.set_title(title)
        ax.legend()

    def generate_horizontal_graph(self, ax):
        """Генерирует горизонтальный график
          Args:
             ax: область в поле для рисования графика
        """
        plt.rcParams['font.size'] = '8'
        ax.set_title("Уровень зарплат по городам")
        cities = self.cities_list_items[0]
        salaries = self.cities_list_items[1]
        labels = [city.replace(' ', '\n').replace('-', '-\n') for city in cities]
        y = np.arange(len(labels))
        ax.barh(y, salaries)
        ax.set_yticks(y, labels=labels, fontsize=6)
        ax.grid(axis='x')
        ax.invert_yaxis()

    def generate_pie(self, ax):
        """Генерирует pie график
           Args:
              ax: область в поле для рисования графика
        """
        plt.rcParams['font.size'] = '6'
        ax.set_title("Доля вакансий по городам")
        labels = self.cities_list_items[3]
        labels.insert(0, "Другие")
        vacancies_share_vals = self.cities_list_items[4]
        vacancies_share_vals.insert(0, 1 - sum(vacancies_share_vals))
        ax.pie(vacancies_share_vals, labels=labels)
        ax.axis('equal')

    def generate_image(self):
        """Метод для создания картинки с графиками.
        """
        fig = plt.figure()
        fig.set_size_inches(10, 7)
        plt.rc('xtick', labelsize=8)
        plt.rc('ytick', labelsize=8)
        ax1 = fig.add_subplot(221)
        ax2 = fig.add_subplot(222)
        ax3 = fig.add_subplot(223)
        ax4 = fig.add_subplot(224)
        self.generate_vertical_graph("Уровень зарплат по годам",
                                     ['средняя з/п', f'з/п {self.profession}'],
                                     self.years_list_items[1],
                                     self.years_list_items[2],
                                     self.years_list_items[0],
                                     ax1)
        self.generate_vertical_graph("Количество вакансий по годам",
                                     ['Количество вакансий', f"Количество вакансий {self.profession}"],
                                     self.years_list_items[3],
                                     self.years_list_items[4],
                                     self.years_list_items[0],
                                     ax2)
        self.generate_horizontal_graph(ax3)
        self.generate_pie(ax4)
        plt.tight_layout()
        plt.savefig('graph.png', dpi=200)

    def format_td(self, list, name):
        """Форматирует данные в виде процентного вида для ячейки таблицы
            Args:
                list (list): excel страница c таблицей
                name (str): название столбца ячеки которого нужно привести в правильный формат
        """
        for cell in list[name][1:]:
            cell.value = str(round(float(cell.value) * 100, 2)).replace('.', ',') + '%'

    def generate_pdf(self):
        """Генерирует pdf файл
        """
        self.generate_image()
        years_sheet, cities_sheet = self.generate_excel()
        self.format_td(cities_sheet, 'E')
        conf_path = r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe'
        props = {'profession': f'{self.profession}',
                 'years_sheet': years_sheet,
                 'cities_sheet': cities_sheet,
                 'cities_sheet_headings': self.cities_list_headings,
                 'years_sheet_headings': self.years_list_headings,
                 'image_file': "graph.png",
                 }
        env = Environment(loader=FileSystemLoader('.'))
        pdf_template = env.get_template("template.html").render(props)
        config = pdfkit.configuration(wkhtmltopdf=conf_path)
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={'enable-local-file-access': True})


def getStatistics():
    file_name = input("Введите название файла: ")
    profession = input("Введите название профессии: ")
    dataset = DataSet(file_name, profession)
    dataset.print_result()
    report(dataset).generate_pdf()


# getStatistics()
# doctest.testmod()
